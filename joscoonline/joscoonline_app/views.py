import math
import re
from datetime import date, datetime
from itertools import zip_longest
from os import truncate
from random import random
from datetime import date, timedelta

import openpyxl
from dateutil.relativedelta import relativedelta
from dateutil.utils import today
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from dateutil.relativedelta import relativedelta
from django.template.defaultfilters import floatformat
from django.template.loader import render_to_string
from openpyxl.workbook import Workbook
from requests.auth import HTTPBasicAuth

from .models import CustomUser
import random
# Create your views here.
from django.db import connection, connections


def home(request):
    q = f"select * from gold_scheme_masters "

    # query = f'''
    # SELECT 1 as Bid,branchname,branchpic, SUM(TotalAmount) as Amt,sum(discount) as discount, sum(tax) as tax,sum(packingcharge) as pk,sum(netamount) as netamt,count(*) as BillCount FROM [TONICOKTM].[TONICOUNIKTM20242025].[dbo].[billmaster] as Ktm,[JSGSRV].[tonicoweb].[dbo].[tonico_erp_branch] as srv where srv.id=1 and Division <>'CHAI' and billdate='{formatted_date} ' group by branchname,branchpic union all SELECT 2 as Bid,branchname,branchpic, SUM(TotalAmount) as Amt ,sum(discount) as discount, sum(tax) as tax,sum(packingcharge) as pk,sum(netamount) as netamt,count(*) as BillCount FROM [TONICOKKD].[TONICOUNIEKM20242025].[dbo].[billmaster] as Ktm,[JSGSRV].[tonicoweb].[dbo].[tonico_erp_branch] as srv where srv.id=2 and billdate='{formatted_date} ' group by branchname,branchpic union all SELECT 3 as Bid,branchname,branchpic, SUM(TotalAmount) as Amt  ,sum(discount) as discount, sum(tax) as tax,sum(packingcharge) as pk,sum(netamount) as netamt,count(*) as BillCount FROM [TONICO-NH].[TONICOUNICKDY20242025].[dbo].[billmaster] as Ktm,[JSGSRV].[tonicoweb].[dbo].[tonico_erp_branch] as srv where srv.id=3 and billdate='{formatted_date} ' group by branchname,branchpic union all SELECT 4 as Bid,branchname,branchpic, SUM(TotalAmount) as Amt  ,sum(discount) as discount, sum(tax) as tax,sum(packingcharge) as pk,sum(netamount) as netamt,count(*) as BillCount FROM [TONICOKTM].[TONICOUNIKTM20242025].[dbo].[billmaster] as Ktm,[JSGSRV].[tonicoweb].[dbo].[tonico_erp_branch] as srv where srv.id=4 and Division ='CHAI' and billdate='{formatted_date} ' group by branchname,branchpic
    # '''
    with connection.cursor() as cursor:
        cursor.execute(q)
        rows = cursor.fetchall()
        data = []
        for row in rows:
            data.append({
                'billdate': row[0],
                # 'billno': row[1],
                # 'tableid': row[2],
                # 'totalqty': row[3],
                # 'totalamount': row[4],
                # 'discount': row[5],
                # 'SGST': (row[6]) / 2,
                # 'CGST': (row[6]) / 2,
                # 'packingcharge': row[7],
                # 'roundoff': row[8],
                # 'netamount': row[9],
                # 'packs': row[10],
                # 'Packagemode': row[11],
                # 'remarks': row[12],
            })
            context={'sdata': data}

    return render(request,"home.html",context)


def login_user(request):
    if(request.method=="POST"):
        uname=request.POST['u']
        pwd=request.POST['p']


        user=authenticate(request,username=uname,password=pwd)

        if (user):
            login(request,user)
            return redirect('scheme')
        else:
            return HttpResponse("Invalid User")

    return render(request,'login.html')


def user_logout(request):
    logout(request)
    return redirect('login')


@login_required
def scheme(request):
    if not request.user.is_authenticated:
        return redirect('login')

    branchid = request.user.branchid
    branchname = "Unknown Branch"  # Default value for branch name

    # Fetch branch name
    query_branch = f"SELECT name FROM branches WHERE id={branchid}"
    with connection.cursor() as cursor:
        cursor.execute(query_branch)
        branch_row = cursor.fetchone()
        if branch_row:
            branchname = branch_row[0]

    # Get `report_type` and date inputs
    report_type = request.POST.get('report', '1')  # Default to '1'
    start_date1 = request.POST.get('start_date1', date.today().strftime('%Y-%m-%d'))
    sdate = request.POST.get('start_date1', date.today().strftime('%d-%m-%Y'))
    end_date1 = request.POST.get('end_date1', date.today().strftime('%Y-%m-%d'))
    edate = request.POST.get('end_date1', date.today().strftime('%d-%m-%Y'))

    # Build SQL query based on report type
    search_query="hello"

    search_query = request.POST.get('search_data', '').strip()
    if search_query !='':
        search_query = request.POST.get('search_data', '').strip()
        QQ = "Joining Report"
        query = f"""
                   SELECT gsm.id, u.name, u.phone, gsm.scheme_id, gsm.erp_scheme_id,
                          sum(gsd.amount) as amount, sum(gsd.gold_weight) as gold_weight ,  ROUND(SUM(gsd.amount)/SUM(gsd.gold_weight), 2) AS gold_rate,  gsm.start_date
                   FROM gold_scheme_masters gsm
                   JOIN gold_scheme_details gsd ON gsm.id = gsd.scheme_master_id
                   JOIN users u ON gsm.user_id = u.id
                   WHERE 
                     gsm.branch_id = {branchid}  and erp_scheme_id ='{ search_query }' group by gsm.id, u.name, u.phone, gsm.scheme_id, gsm.erp_scheme_id,gsm.start_date;
               """
    else:
        if report_type == '1':
            QQ = "Joining Report"
            query = f"""
                SELECT gsm.id, u.name, u.phone, gsm.scheme_id, gsm.erp_scheme_id,
                       SUM(gsd.amount) AS amount,
                       TRUNCATE(SUM(gsd.gold_weight), 3) AS gold_weight,
                       ROUND(SUM(gsd.amount)/SUM(gsd.gold_weight), 2) AS gold_rate,
                       gsm.start_date
                FROM gold_scheme_masters gsm
                JOIN gold_scheme_details gsd ON gsm.id = gsd.scheme_master_id
                JOIN users u ON gsm.user_id = u.id
                WHERE gsm.start_date BETWEEN '{start_date1}' AND '{end_date1}'
                  AND gsm.branch_id = {branchid}
                GROUP BY gsm.id, u.name, u.phone, gsm.scheme_id, gsm.erp_scheme_id, gsm.start_date;
            """
        elif report_type == '2':
            QQ = "Payment Report"
            query = f"""
            SELECT gsm.id, u.name, u.phone, gsm.scheme_id, gsm.erp_scheme_id,
                  
                   gsd.amount, gsd.gold_weight, gsd.gold_rate, gsd.current_date,gsd.raz_order_id,gsd.transaction_ID
            FROM gold_scheme_masters gsm
            JOIN gold_scheme_details gsd ON gsm.id = gsd.scheme_master_id
            JOIN users u ON gsm.user_id = u.id
            WHERE gsd.current_date BETWEEN '{start_date1}' AND '{end_date1}'
              AND gsm.branch_id = {branchid};
            """
        elif report_type == '3':
                QQ = "Closed Report"
                query = f"""
                   SELECT gsm.id, u.name, u.phone, gsm.scheme_id, gsm.erp_scheme_id,
                   
                   sum(gsd.amount) as amount, TRUNCATE(sum(gsd.gold_weight),3) as gold_weight , ROUND(SUM(gsd.amount)/SUM(gsd.gold_weight), 2) AS gold_rate, gsm.closing_date
            FROM gold_scheme_masters gsm
                    JOIN gold_scheme_details gsd ON gsm.id = gsd.scheme_master_id
                    JOIN users u ON gsm.user_id = u.id
                    WHERE gsm.closing_date BETWEEN '{start_date1}' AND '{end_date1}'
                      AND gsm.branch_id = {branchid} group by gsm.id, u.name, u.phone, gsm.scheme_id, gsm.erp_scheme_id , gsm.closing_date; """
        else:
                   pass

    # Execute the query and process results
    with connection.cursor() as cursor:

        cursor.execute(query)
        rows = cursor.fetchall()

    data = []

    for row in rows:
        order_id = row[9] if len(row) > 10 else 0
        payment_id = row[10] if len(row) > 10 else 0
        is_order = str(order_id).startswith("order_")

        # Fetch data only once to avoid redundant calls
        order_data = fetch_orders(order_id) if is_order else {}
        payment_data = fetch_payments(order_id) if is_order else {}
        settlement_data = fetch_settlement_by_order(order_id) if is_order else [{}]

        # Add processed row to data
        data.append({
            'id': row[0],
            'name': row[1],
            'phone': row[2],
            'scheme_id': row[3],
            'erp_scheme_id': row[4],
            'payment_method': row[4],
            'amount': row[5],
            'goldweight': row[6],
            'goldrate': row[7],
            'start_date': row[8].strftime('%d/%m/%Y'),
            'order_id': order_id,
            'payment_id': payment_id,
            'order_status': order_data.get('status', '').upper(),
            'amount_paid': int(order_data.get('amount_paid', 0) / 100),
            'pay_status': payment_data.get('status', '').upper(),
            'method': payment_data.get('method', '').upper(),
            'vpa': payment_data.get('vpa', ''),
            'email': payment_data.get('email', ''),
            'contact': payment_data.get('contact', ''),
            'rrn': payment_data.get('acquirer_data', {}).get('rrn', ''),
            'upi_transaction_id':payment_data.get('acquirer_data', {}).get('upi_transaction_id', ''),
            'settlement_utr': settlement_data[0].get('settlement_utr', ''),
            'credit': settlement_data[0].get('credit', 0)/100,
            'fee': (settlement_data[0].get('fee', 0) / 100)-(settlement_data[0].get('tax', 0) / 100),
            'tax': settlement_data[0].get('tax', 0) / 100
        })



    # Calculate totals
    total_amt =  sum(int(float(item['amount'])) for item in data)
    total_wt = sum(float(item['goldweight']) for item in data)

    # Prepare context and render
    context = {
        "sdata": data,
        "branchid": branchid,
        "totalamt": total_amt,
        "totalwt": float(f"{total_wt:.3f}"),
        "sdate": sdate,
        "edate": edate,
        "branch": branchname,
        "rtype": QQ,
        "searchquery":search_query,
    }


    return render(request, 'scheme.html', context)


def chitdetails(request, id):
    schemeid=0
    if request.user.is_authenticated:
        branchid = request.user.branchid

        # Query for gold scheme details
        query = f"""
            SELECT 
                CASE
                    WHEN gsd.payment_method = 1 THEN 'Cash'
                    WHEN gsd.payment_method = 2 THEN 'UPI'
                    WHEN gsd.payment_method = 3 THEN 'Card'
                    WHEN gsd.payment_method = 4 THEN 'Netbanking'
                    ELSE 'Unknown'
                END AS payment_method,
                gsd.amount,
                gsd.gold_weight,
                gsd.gold_rate,
                gsd.current_date,
                gsm.id
            FROM 
                gold_scheme_masters gsm
            JOIN 
                gold_scheme_details gsd ON gsm.id = gsd.scheme_master_id
            JOIN 
                users u ON gsm.user_id = u.id
            WHERE 
                erp_scheme_id = '{id}' AND gsm.branch_id = {branchid};
        """

        with connection.cursor() as cursor2:
            cursor2.execute(query)
            rows = cursor2.fetchall()
            data = [
                {
                    'transid': row[0],
                    'amount': row[1],
                    'goldweight': row[2],
                    'goldrate': row[3],
                    'start_date': row[4].strftime('%d/%m/%Y') ,
                    'schemeid=': row[5]
                }
                for row in rows
            ]
        if rows:
          schemeid = rows[-1][5]
        # Query for user details and nominee info
        query1 = f"""
            SELECT 
                kyc.name,
                u.phone,
                kyc.address,
                gsm.start_date,
                gsm.scheme_end_date,
                gsm.status,
                gsm.closing_date,
                gsm.invoice_number,
                kyc.type,
                gsm.scheme_peroid
               
            FROM 
                gold_scheme_masters gsm
            JOIN 
                users u ON gsm.user_id = u.id
            LEFT JOIN 
                user_kycs kyc ON u.id = kyc.user_id
            WHERE 
                gsm.erp_scheme_id = '{id}' AND gsm.branch_id = {branchid}
                AND u.name <> ''
                AND (kyc.user_id IS NULL OR kyc.name <> '');
        """

        with connection.cursor() as cursor3:
            cursor3.execute(query1)
            rows1 = cursor3.fetchall()
            data1 = []
            today = datetime.today()

            for row1 in rows1:
                start_date = row1[3]
                months_diff = (
                        (today.year - start_date.year) * 12 + today.month - start_date.month
                ) if start_date else 0


                address_trimmed = row1[2].split('[')[0] if row1[2] and '[' in row1[2] else row1[2] or ""
                if row1[2] and len(row1[2]) > 0:
                    address_trimmed = row1[2].split('[')[0] if '[' in row1[2] else row1[2]
                    match = re.search(r'\[.*?\]', row1[2])
                    nominee_info = match.group(0) if match else " "  # Extract the nominee info

                else:
                    address_trimmed = ""
                    nominee_info = ""

                data1.append({
                    'name': row1[0],
                    'phone': row1[1],
                    'address': address_trimmed,
                    'start_date': row1[3],
                    'end_date': row1[4],
                    'status': row1[5],
                    'closingdate': row1[6],
                    'invoice': row1[7],
                    'type': row1[8],
                    'period': row1[9],
                    'monthdif': months_diff,
                    'nomineeinfo': nominee_info,
                })

        # Totals
        total_amt = sum(int(float(item.get('amount'))) for item in data)
        total_wt = sum(float(item['goldweight']) for item in data)

        # Context for rendering
        context = {
            "sdata": data,
            "totalamt": total_amt,
            "totalwt": float(f"{total_wt:.3f}"),
            "mdata": data1,"schemeID":schemeid,
        }

        return render(request, "chitdetails.html", context)

    # If not authenticated, redirect or show an error page
    return HttpResponse("Unauthorized", status=401)

def addpayment(request):
    if (request.method == "POST"):
        schemeID= request.POST.get('schemeID')
        dt = request.POST.get('date')
        grate=int(request.POST.get('gold_rate'))
        amt=int(request.POST.get('amount'))
        pm=request.POST.get('payment_method')
        wt= amt/grate

        dt_value = datetime.strptime(dt, "%Y-%m-%d")  # Convert string to datetime if needed

        # Add 1 month to the date
        new_date = dt_value + relativedelta(months=1)

        # Format the date to 'yyyy-MM-dd'
        formatted_date = new_date.strftime("%Y-%m-%d")

        query=f""" insert into gold_scheme_details(scheme_master_id,amount,gold_weight,gold_rate,status,payment_status,raz_order_id,transaction_id,next_due,`current_date`,payment_method,payment_mode) values({ schemeID} ,{amt} ,{math.trunc(wt*1000)/1000},{grate} ,1,1,'BANK','BANK','{formatted_date }','{  dt_value}',{pm},1)     """
        with connection.cursor() as cursor:
            cursor.execute(query)
            # Commit the transaction if needed (unnecessary if you have autocommit enabled)
            connection.commit()

        query1=f""" select * from gold_scheme_masters where id={schemeID}"""
        with connection.cursor() as cursor1:
            cursor1.execute(query1)
            result = cursor1.fetchone()  # Get the first row (if there is one)
            amount1=int(result[4])
            gwt1=result[5]
            ERPID=result[3]

            gamt=int(amount1)+int(amt)
            gwt=math.trunc((float(gwt1)+float(wt))*1000)/1000

        query1 = f""" update gold_scheme_masters  set amount_accumulated ={gamt},gold_accumulated={gwt},total_amount_accumulated={gamt},total_gold_weight={gwt} where id={schemeID}"""
        with connection.cursor() as cursor2:
            cursor2.execute(query1)
   # chitdetails(request,  ERPID)
        return HttpResponse("""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Thank You</title>
        <!-- Bootstrap CSS -->
        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
    </head>
    <body>
        <div class="container d-flex align-items-center justify-content-center vh-100">
            <div class="text-center">
                <div class="mb-4">
                    <img src="https://via.placeholder.com/150" alt="Thank You" class="img-fluid rounded-circle">
                </div>
                <h1 class="display-4 text-success">Thank You!</h1>
                <p class="lead">Your payment has been processed successfully.</p>
                <p class="text-muted">We appreciate your trust in our services.</p>
                <a href="/" class="btn btn-primary mt-3" onclick="closeWindow()">Close</a>
            </div>
        </div>

        <!-- Bootstrap JS -->
        <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
    </body>
    </html>
    <script>
    function closeWindow() {
        window.open('', '_self'); // Ensure compatibility
        window.close();
    }
</script>
    """)

# def get_gold_rate(request):
#         print(request.GET.get('todate') + "  test     ")
#         selected_date = request.GET.get('todate')
#         try:
#             query=f""" select rate from gold_rate_logs where `current_date`='{selected_date}' and gold_purity=4     """
#             with connection.cursor() as cursor3:
#                 cursor3.execute(query1)
#                 rows1 = cursor3.fetchone()
#                 gold_rate = rows1[0]
#                 print(gold_rate + "      Sucess")
#                 return JsonResponse({'success': True, 'gold_rate': str(gold_rate)})
#         except:
#                 print("     Error")
#                 return JsonResponse({'success': False, 'gold_rate': str(0)})
def addnewscheme(request):
    context = {}
    if request.method == 'POST':
        today = datetime.today()
        formatted_date = today .strftime("%Y-%m-%d")
        query = f""" SELECT rate FROM gold_rate_logs WHERE `current_date`='{formatted_date}' and gold_purity=4 """ # Use parameterized query to prevent SQL injection

        with connection.cursor() as cursor:
            cursor.execute(query)
            result2 = cursor.fetchone()  # Get the first matching row
            if result2:
                context['gold_rate'] = result2[0]
            else:
                context['gold_rate'] = 0

        phone = request.POST.get('phone')
        name = request.POST.get('name')
        email = request.POST.get('email')
        context['name'] =name
        context['phone'] =phone
        context['email'] = email

        query = "SELECT * FROM users WHERE phone = %s"  # Use parameterized query to prevent SQL injection

        with connection.cursor() as cursor:
            cursor.execute(query, [phone])
            result = cursor.fetchone()  # Get the first matching row

        if result:
            context['ID'] = result[0]  # Assuming the address is in the first column

            query3 = "SELECT erp_scheme_id FROM gold_scheme_masters WHERE user_id = %s"

            with connection.cursor() as cursor5:
                cursor5.execute(query3, [result[0]])  # Pass the parameter as a list
                result5 = cursor5.fetchall() # Get the first matching row
                data = [
                    {
                        'schemes': row[0]

                    }
                    for row in result5
                ]
                context['schemes'] = data

            query = f"""  SELECT * FROM user_kycs WHERE User_id ={result[0]} """  # Use parameterized query to prevent SQL injection

            with connection.cursor() as cursor:
                cursor.execute(query)
                result1 = cursor.fetchone()  # Get the first matching row
                if result1:
                       context['address'] = result1[3]
                else:
                    context['address'] = "None"
            context['exists'] = True
        else:

            query = f"""  insert into users (name,email,country_code,phone,status,is_complete) values('{name}','{email}','91','{phone}',1,1) """  # Use parameterized query to prevent SQL injection

            with connection.cursor() as cursor:
                cursor.execute(query)
            query = "SELECT * FROM users WHERE phone = %s"  # Use parameterized query to prevent SQL injection

            with connection.cursor() as cursor:
                cursor.execute(query, [phone])
                result = cursor.fetchone()  # Get the first matching row
            context['ID'] = result[0]
            context['phone'] = phone  # Keep the phone number in the form for user reference

    return render(request, 'addnewscheme.html', context)


def addnewschemedetails(request):
    if request.method == 'POST':
        try:
            # Get data from POST request
            user_id = request.POST.get('user_id')
            address = request.POST.get('address')
            document = request.POST.get('document')  # Not used, remove if unnecessary
            relation = request.POST.get('relation')  # Not used, remove if unnecessary
            types = request.POST.get('type')
            dates = request.POST.get('date')
            gold_rate = request.POST.get('gold_rate')
            amount = request.POST.get('amount')
            name = request.POST.get('name1')
            payment_method = request.POST.get('payment_method')
            branchid = request.user.branchid

            # Validate required fields
            if not (user_id and address and types and dates and gold_rate and amount and name and payment_method):
                return HttpResponse("Missing required fields.", status=400)

            # Calculate derived values
            grate = int(gold_rate)
            amt = int(amount)
            gtt = math.trunc((amt / grate) * 1000) / 1000
            wt = gtt

            # Generate scheme_id and transaction_id
            scheme_id = generate_unique_scheme_id()
            tid = f"{payment_method}_{random.randint(1, 100000)}"

            # Calculate dates
            start_date = datetime.strptime(dates, "%Y-%m-%d")
            next_due = start_date + relativedelta(months=1)
            scheme_end_date = start_date + relativedelta(months=10)

            # Format dates for SQL
            start_date_str = start_date.strftime('%Y-%m-%d')
            next_due_str = next_due.strftime('%Y-%m-%d')
            scheme_end_date_str = scheme_end_date.strftime('%Y-%m-%d')

            # Insert into user_kycs
            if not searchrecord(user_id):

                user_kyc_query = """
                INSERT INTO user_kycs (user_id, name, address, type, status)
                VALUES (%s, %s, %s, %s, %s)
            """
            with connection.cursor() as cursor:
                cursor.execute(user_kyc_query, [user_id, name, address, types, 0])

            # Insert into gold_scheme_masters
            gold_scheme_master_query = """
                INSERT INTO gold_scheme_masters 
                (user_id, scheme_id, amount_accumulated, gold_accumulated, last_reduced, total_gold_weight, 
                 order_id, total_amount_accumulated, purchased_amount, branch_id, refund, discount_percentage, 
                 status, scheme_peroid, start_date, next_due, scheme_end_date, erp_scheme_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            with connection.cursor() as cursor:
                cursor.execute(gold_scheme_master_query, [
                    user_id, scheme_id, amt, wt, 0, wt, 0, amt, 0, branchid, 0, 3.5, 1, 10,
                    start_date_str, next_due_str, scheme_end_date_str, ''
                ])

            # Get the newly inserted gold_scheme_masters ID
            get_scheme_id_query = """
                SELECT id FROM gold_scheme_masters WHERE user_id=%s AND scheme_id=%s
            """
            with connection.cursor() as cursor:
                cursor.execute(get_scheme_id_query, [user_id, scheme_id])
                result = cursor.fetchone()
                if not result:
                    return HttpResponse("Failed to retrieve scheme ID.", status=500)

                scheme_master_id = result[0]

            # Insert into gold_scheme_details
            gold_scheme_details_query = """
                INSERT INTO gold_scheme_details 
                (scheme_master_id, amount, gold_weight, gold_rate, status, payment_status, 
                 raz_order_id, transaction_id, next_due, `current_date`, payment_mode, payment_method)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            with connection.cursor() as cursor:
                cursor.execute(gold_scheme_details_query, [
                    scheme_master_id, amt, wt, grate, 1, 1, payment_method, tid,
                    next_due_str, start_date_str, 1, payment_method
                ])

            # Return success page
            return HttpResponse("""
                <!DOCTYPE html>
                <html lang="en">
                <head>
                    <meta charset="UTF-8">
                    <meta name="viewport" content="width=device-width, initial-scale=1.0">
                    <title>Thank You</title>
                    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
                </head>
                <body>
                    <div class="container d-flex align-items-center justify-content-center vh-100">
                        <div class="text-center">
                            <div class="mb-4">
                                <img src="https://via.placeholder.com/150" alt="Thank You" class="img-fluid rounded-circle">
                            </div>
                            <h1 class="display-4 text-success">Thank You!</h1>
                            <p class="lead">New Scheme processed successfully.</p>
                            <p class="text-muted">We appreciate your trust in our services.</p>
                            <a href="/" class="btn btn-primary mt-3">Close</a>
                        </div>
                    </div>
                </body>
                </html>
            """)

        except Exception as e:
            return HttpResponse(f"An error occurred: {str(e)}", status=500)

def generate_unique_scheme_id():
    scheme_id = ""
    is_unique = False

    while not is_unique:
        # Generate a random scheme ID in the format "JS_O_XXXXXXXX" where X is a random 8-digit number
        scheme_id = f"JS_O_{random.randint(0, 99999999):08d}"

        # Check if the generated scheme ID already exists in the database
        query = "SELECT COUNT(*) FROM gold_scheme_masters WHERE scheme_id = %s"
        with connection.cursor() as cursor:
            cursor.execute(query, [scheme_id])
            count = cursor.fetchone()[0]

        # If count is 0, the scheme ID is unique
        if count == 0:
            is_unique = True

    return scheme_id



def searchrecord(table,condition):
    query = f"""SELECT COUNT(*) FROM {table} WHERE {condition}"""
    is_unique = True
    with connection.cursor() as cursor:
        cursor.execute(query)
        count = cursor.fetchone()[0]
    if count == 0:
        is_unique = False

    return(is_unique)


def findrecord(table,condition,value):
    query = f"""SELECT {value} FROM {table} WHERE {condition}"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        result = cursor.fetchone()
        if result:
            # If the result is a tuple (e.g., (id,))
            return result[0] if isinstance(result, tuple) else result
        else:
            return 0






@login_required
def indsoftintegration(request):
    context = {}

    branchid = request.user.branchid

    start_date1 = request.POST.get('start_date1', date.today().strftime('%m-%d-%Y'))
    sdate = request.POST.get('start_date1', date.today().strftime('%d-%m-%Y'))
    end_date1 = request.POST.get('end_date1', date.today().strftime('%m-%d-%Y'))
    edate = request.POST.get('end_date1', date.today().strftime('%d-%m-%Y'))

    try:
        cn = str(branchid)  # Assuming 'ekm' is the connection alias
        with connections[cn].cursor() as cursor:


            # Create the table if it doesn't exist
            query = """ 
                IF NOT EXISTS (SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = 'CHITRCPTID')
                CREATE TABLE CHITRCPTID (
                    id INT 
                );
                """
            cursor.execute(query)

            # Fetch data from the database
            query = f"""
                SELECT DISTINCT cm.chit_key, chit_no, cm.cust_name, 
                    ccm.Address + ', ' + ccm.Address1 + ', ' + ccm.Address2 + 
                    ' [ Nominee : ' + cm.Nominee + '( ' + cm.Nrelation + ' )]' AS Address, 
                    CASE 
                        WHEN cm.Mobile = '' THEN ccm.phone_no 
                        ELSE cm.Mobile 
                    END AS Phone_no,
                    cm.Cust_Code
                FROM chitmast cm 
                JOIN CHITCUSTMAST ccm ON cm.ccust_key = ccm.CCust_key
                JOIN chitrcpt crp ON crp.chit_key = cm.chit_key
                WHERE cm.close_flag = 'N' 
                AND CONVERT(DATETIME, vch_date, 103) BETWEEN '{start_date1}' AND '{end_date1}' and crp.serial_no not in (select id from CHITRCPTID where id is not null)
                ORDER BY cm.chit_key
                """
            cursor.execute(query)
            result = cursor.fetchall()

            data = [
                {
                    'chit_key': row[0],
                    'chit_no': row[1],
                    'cust_name': row[2],
                    'address': row[3],
                    'phone': row[4],
                    'cust_code': row[5]
                }
                for row in result
            ]

            context = {'sdata': data}
            return render(request, 'indsoftintegration.html', context)

    except Exception as e:

        # Ensure that the view returns an HttpResponse even in case of an exception
        return HttpResponse("An error occurred while fetching data." + str(e))



@login_required
def webupdate(request, cust_code=None, chit_key=None):

    if request.POST:
        branchid = request.user.branchid
        cn = str(branchid)  # Assuming 'ekm' is the connection alias

        query = """
        SELECT 
            cm.chit_key,
            cm.chit_no,
            cm.cust_name,
            cm.email,
            ccm.Address + ', ' + ccm.Address1 + ', ' + ccm.Address2 + ' [ Nominee :' + cm.Nominee + '( ' + cm.Nrelation + ' )]' AS Address, 
            CASE 
                WHEN cm.Mobile = '' THEN ccm.phone_no 
                ELSE cm.Mobile 
            END AS Phone_no,
            cm.Cust_Code,
            cm.inst_no,
            cm.doj
        FROM chitmast cm 
        INNER JOIN CHITCUSTMAST ccm ON cm.ccust_key = ccm.CCust_key 
        WHERE cm.cust_code = %s
        """


        with connections[cn].cursor() as cursor1:
            cursor1.execute(query, [cust_code])  # Use parameterized query to prevent SQL injection
            result = cursor1.fetchone()
            print(cust_code)
            if result:
                start_date = datetime.strptime(result[8], "%d/%m/%Y")
                next_due = start_date + relativedelta(months=1)
                scheme_end_date = start_date + relativedelta(months=10)

                # Format dates for SQL
                start_date_str = start_date.strftime('%Y-%m-%d')
                next_due_str = next_due.strftime('%Y-%m-%d')
                scheme_end_date_str = scheme_end_date.strftime('%Y-%m-%d')

                scheme_id = f"JS_O_{random.randint(0, 99999999):08d}"

                # Check if the record exists in gold_scheme_masters
                if not searchrecord("gold_scheme_masters", f"erp_scheme_id='{cust_code}' and branch_id={branchid}"):

                    # Insert into users table
                    query = """
                    INSERT INTO users (name, email, country_code, phone, status, is_complete) 
                    VALUES (%s, %s, '91', %s, 1, 1)
                    """
                    with connection.cursor() as cursor:
                        cursor.execute(query, [result[2], result[3], result[5]])

                    user_id = findrecord("users", f"phone='{result[5]}' and name='{result[2]}'", "id")

                    # Insert into user_kycs table
                    query = """
                    INSERT INTO user_kycs (user_id, name, address, type, status) 
                    VALUES (%s, %s, %s, 'Aadhar Card', 1)
                    """
                    with connection.cursor() as cursor:
                        cursor.execute(query, [user_id, result[2], result[4]])

                    # Insert into gold_scheme_masters
                    query = f"""
                    INSERT INTO gold_scheme_masters 
                    (user_id, scheme_id, amount_accumulated, gold_accumulated, last_reduced, total_gold_weight, order_id, total_amount_accumulated, purchased_amount, branch_id, refund, discount_percentage, status, scheme_peroid, start_date, next_due, scheme_end_date, erp_scheme_id) 
                    VALUES ({user_id}, '{scheme_id}', 0, 0, 0, 0, 0, 0, 0, {branchid}, 0, 3.5, 1, {result[7]}, '{start_date_str}', '{next_due_str}', '{scheme_end_date_str}', '{cust_code}')                    """
                    with connection.cursor() as cursor:
                        cursor.execute(query)

                    # Find scheme master id
                    scheme_masterid = findrecord("gold_scheme_masters",
                                                 f"erp_scheme_id='{cust_code}' and branch_id={branchid}", "id")

                    # Insert into gold_scheme_details for each record in chitrcpt
                    query = """
                    SELECT 
                        vch_no, vch_date, Amount, conv_rate, gold_wt, rcpt_no,
                        CASE 
                            WHEN trx_type = 'C' THEN 1 
                            WHEN trx_type = 'A' THEN 
                                CASE 
                                    WHEN cheque_no = '' THEN 2 
                                    ELSE 4 
                                END
                            WHEN trx_type = 'B' THEN 4 
                            ELSE 4 
                        END AS Trx, serial_no 
                    FROM chitrcpt 
                    WHERE chit_key = %s
                    """
                    cursor1.execute(query, [chit_key])
                    rows = cursor1.fetchall()
                    amt = 0
                    gwt = 0

                    # Insert the records into gold_scheme_details
                    with connection.cursor() as cursor:
                        for row1 in rows:
                            start_date = datetime.strptime(row1[1], "%d/%m/%Y")
                            next_due_date = start_date + relativedelta(months=1)
                            next_due_str = next_due_date.strftime("%Y-%m-%d")

                            cdate = datetime.strptime(row1[1], "%d/%m/%Y")

                            cdt= cdate.strftime("%Y-%m-%d")




                            query = f"""
                            INSERT INTO gold_scheme_details 
                            (scheme_master_id, amount, gold_weight, gold_rate, status, payment_status, raz_order_id, transaction_id, next_due, `current_date`, payment_mode, payment_method) 
                            VALUES ({scheme_masterid}, {row1[2]}, {row1[4]}, {row1[3]}, 1, 1, 'BANK','{row1[5]}', '{next_due_str}', '{cdt}',  1, {row1[6]})
                            """
                            cursor.execute(query)

                            query = f""" select id from CHITRCPTID where id={row1[7]} """
                            with connections[cn].cursor() as cursor1:
                                cursor1.execute(query)
                                result = cursor1.fetchone()
                            if not result:
                                query = f"""insert into  CHITRCPTID (id) values( {row1[7]})  """

                            with connections[cn].cursor() as cursor1:
                                cursor1.execute(query)

                            amt += row1[2]
                            gwt += row1[4]

                    # Update gold_scheme_masters with accumulated amounts
                    # print(next_due_str)
                    # print([amt, int(gwt * 1000) / 1000, int(gwt * 1000) / 1000, amt, cust_code, branchid, next_due_str])
                    query = f"""
                    UPDATE gold_scheme_masters 
                    SET amount_accumulated = {amt}, gold_accumulated = {int(gwt * 1000) / 1000}, total_gold_weight = {int(gwt * 1000) / 1000}, total_amount_accumulated = {amt} ,next_due='{next_due_str}'
                    WHERE erp_scheme_id ='{cust_code}' AND branch_id = {branchid}
                    """
                    with connection.cursor() as cursor:
                        cursor.execute(query)
                    return HttpResponse("""
                                                           <!DOCTYPE html>
                                                           <html lang="en">
                                                           <head>
                                                               <meta charset="UTF-8">
                                                               <meta name="viewport" content="width=device-width, initial-scale=1.0">
                                                               <title>Thank You</title>
                                                               <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
                                                           </head>
                                                           <body>
                                                               <div class="container d-flex align-items-center justify-content-center vh-100">
                                                                   <div class="text-center">
                                                                       <div class="mb-4">
                                                                           <img src="https://via.placeholder.com/150" alt="Thank You" class="img-fluid rounded-circle">
                                                                       </div>
                                                                       <h1 class="display-4 text-success">Thank You!</h1>
                                                                       <p class="lead">All Scheme closed  successfully for Today.</p>
                                                                       <p class="text-muted">We appreciate your trust in our services.</p>
                                                                       <a href="javascript:window.history.back();" class="btn btn-primary mt-3">Close</a>
                                                                   </div>
                                                               </div>
                                                           </body>
                                                           </html>
                                                       """)
                else:
                    # Existing scheme master
                    scheme_masterid = findrecord("gold_scheme_masters",
                                                 f"erp_scheme_id='{cust_code}' and branch_id={branchid}", "id")

                    # Delete previous details
                    query = "DELETE FROM gold_scheme_details WHERE scheme_master_id = %s and raz_order_id NOT LIKE %s and transaction_id NOT LIKE %s"
                    with connection.cursor() as cursor:
                        cursor.execute(query, (scheme_masterid,'order_%','pay_%'))

                    # Insert new details
                    query = """
                    SELECT 
                        vch_no, vch_date, chitrcpt.Amount, chitrcpt.conv_rate, gold_wt, chitrcpt.rcpt_no,
                        CASE 
                            WHEN chitrcpt.trx_type = 'C' THEN 1 
                            WHEN chitrcpt.trx_type = 'A' THEN 
                                CASE 
                                    WHEN cheque_no = '' THEN 2 
                                    ELSE 4 
                                END
                            WHEN chitrcpt.trx_type = 'B' THEN 4 
                            ELSE 4 
                        END AS Trx ,chitrcpt.serial_no
                    FROM chitrcpt,MCHITRCPTMAST 
                    WHERE chit_key = %s and chitrcpt.MSer_No =MCHITRCPTMAST.Serial_No and rcpttype=0 
                    """
                    cursor1.execute(query, [chit_key])
                    rows = cursor1.fetchall()
                    amt = findrecord("gold_scheme_details",f"scheme_master_id = { scheme_masterid}","sum(amount)")
                    gwt = findrecord("gold_scheme_details",f"scheme_master_id = { scheme_masterid}","sum(gold_weight)")
                    if amt is None: amt=0
                    if gwt is None: gwt=0
                    # Insert the records into gold_scheme_details
                    with connection.cursor() as cursor:
                        for row1 in rows:
                            start_date = datetime.strptime(row1[1], "%d/%m/%Y")
                            next_due_date = start_date + relativedelta(months=1)
                            next_due_str = next_due_date.strftime("%Y-%m-%d")

                            cdate = datetime.strptime(row1[1], "%d/%m/%Y")

                            cdt = cdate.strftime("%Y-%m-%d")

                            query = f"""
                                                       INSERT INTO gold_scheme_details 
                                                       (scheme_master_id, amount, gold_weight, gold_rate, status, payment_status, raz_order_id, transaction_id, next_due, `current_date`, payment_mode, payment_method) 
                                                       VALUES ({scheme_masterid}, {row1[2]}, {row1[4]}, {row1[3]}, 1, 1, 'BANK','{row1[5]}', '{next_due_str}', '{cdt}',  1, {row1[6]})
                                                       """
                            cursor.execute(query)

                            # Update CHITRCPTID

                            query=f"""select id from CHITRCPTID where id={row1[7]} """
                            with connections[cn].cursor() as cursor1:
                                cursor1.execute(query)
                                result=cursor1.fetchone()
                            if not result:
                             query = f"""insert into  CHITRCPTID (id) values( {row1[7]})  """

                            with connections[cn].cursor() as cursor1:
                                cursor1.execute(query)

                            amt += row1[2]
                            gwt += row1[4]

                            # Update gold_scheme_masters with accumulated amounts
                        query = f"""
                                            UPDATE gold_scheme_masters 
                                            SET amount_accumulated = {amt}, gold_accumulated = {int(gwt * 1000) / 1000}, total_gold_weight = {int(gwt * 1000) / 1000}, total_amount_accumulated = {amt} ,next_due='{next_due_str}'
                                            WHERE erp_scheme_id ='{cust_code}' AND branch_id = {branchid}
                                            """

                        with connection.cursor() as cursor:
                            cursor.execute(query)

                        return HttpResponse("""
                                        <!DOCTYPE html>
                                        <html lang="en">
                                        <head>
                                            <meta charset="UTF-8">
                                            <meta name="viewport" content="width=device-width, initial-scale=1.0">
                                            <title>Thank You</title>
                                            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
                                        </head>
                                        <body>
                                            <div class="container d-flex align-items-center justify-content-center vh-100">
                                                <div class="text-center">
                                                    <div class="mb-4">
                                                        <img src="https://via.placeholder.com/150" alt="Thank You" class="img-fluid rounded-circle">
                                                    </div>
                                                    <h1 class="display-4 text-success">Thank You!</h1>
                                                    <p class="lead">All Scheme closed  successfully for Today.</p>
                                                    <p class="text-muted">We appreciate your trust in our services.</p>
                                                    <a href="javascript:window.history.back(); " class="btn btn-primary mt-3">Close</a>
                                                </div>
                                            </div>
                                        </body>
                                        </html>
                                    """)
@login_required
def adduser(request):
    if request.method == "POST":
        # Handle form submission
        username = request.POST.get('username')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        branch_id = request.POST.get('branch')
        query = f""" select id,name from branches order by id"""
        with connection.cursor() as cursor:
            cursor.execute(query)
            tempbranch = cursor.fetchall()
        branches = [
            {
                'id': row[0],
                'name': row[1],
                            }
            for row in tempbranch
        ]




        if password != confirm_password:
            return render(request, 'adduser.html', {
                'branches': branches,
                'error': 'Passwords do not match.',
            })
        else:
            u = CustomUser.objects.create_user(username=username, password=password, first_name=first_name,
                                               last_name=last_name ,branchid=branch_id )
            u.save()
        # Save the user data or perform further processing here
        # Assuming you have a User model to save this data
        # User.objects.create(username=username, first_name=first_name, last_name=last_name, password=password, branch_id=branch_id)

        return render(request, 'adduser.html', {
            'branches': branches,
            'success': 'User registered successfully.',
        })
    else:
        # Render the form with branches
        query = f""" select id,name from branches order by id"""
        with connection.cursor() as cursor:
          cursor.execute(query)
          tempbranch = cursor.fetchall()
          branches = [
              {
                  'id': row[0],
                  'name': row[1],
              }
              for row in tempbranch
          ]

        # branches = cursor.fetchall()
        return render(request, 'adduser.html', {'branches': branches})


        #return render(request, 'adduser.html')

@login_required
def closechit(request):
    if request.method == 'POST':
        branchid = request.user.branchid
        cn = str(branchid)  # Assuming 'ekm' is the connection alias
        start_date1 = request.POST.get('start_date1', datetime.today().strftime('%m-%d-%Y'))
        sdate = request.POST.get('start_date2', datetime.today().strftime('%m/%d/%Y'))
        end_date1 = request.POST.get('end_date1', datetime.today().strftime('%m-%d-%Y'))
        edate = request.POST.get('end_date2', datetime.today().strftime('%m/%d/%Y'))

        sdate = datetime.strptime( sdate, "%Y-%m-%d")

        # Format to MM/DD/YYYY
        sdate = sdate.strftime("%m/%d/%Y")

        edate = datetime.strptime(edate, "%Y-%m-%d")

        # Format to MM/DD/YYYY
        edate = edate.strftime("%m/%d/%Y")




        query = f"""
            SELECT CUST_CODE, CHITMAST.Chit_Key, Payed, CASH, CHKAMT, CHKNO, Ref_NO,ref_date
            FROM CHITMAST, CHITCLOSE
            WHERE CHITMAST.Chit_Key = CHITCLOSE.Chit_Key
            AND convert(datetime,Ref_Date,103) BETWEEN   '{sdate}' AND '{edate}'
        """
        MN=0
        try:
            with connections[cn].cursor() as cursor1:
                cursor1.execute(query)
                result = cursor1.fetchall()

                if result:
                    # Process each row from the result
                    for row in result:
                        # Insert and delete logic as needed
                        scheme_masterid = findrecord("gold_scheme_masters",
                                                     f"erp_scheme_id='{row[0]}' and branch_id={branchid}", "id")

                        delete_query = "DELETE FROM gold_scheme_details WHERE scheme_master_id = %s and raz_order_id NOT LIKE %s and transaction_id NOT LIKE %s"
                        with connection.cursor() as cursor:
                            cursor.execute(delete_query,  (scheme_masterid, 'order_%', 'pay_%'))

                        # Insert new records
                        insert_query = """
                            SELECT 
                        vch_no, vch_date, chitrcpt.Amount, chitrcpt.conv_rate, gold_wt, chitrcpt.rcpt_no,
                        CASE 
                            WHEN chitrcpt.trx_type = 'C' THEN 1 
                            WHEN chitrcpt.trx_type = 'A' THEN 
                                CASE 
                                    WHEN cheque_no = '' THEN 2 
                                    ELSE 4 
                                END
                            WHEN chitrcpt.trx_type = 'B' THEN 4 
                            ELSE 4 
                        END AS Trx ,chitrcpt.serial_no
                    FROM chitrcpt,MCHITRCPTMAST 
                    WHERE chit_key = %s and chitrcpt.MSer_No =MCHITRCPTMAST.Serial_No and rcpttype=0 
                        """
                        with connections[cn].cursor() as cursor1:
                            cursor1.execute(insert_query, (row[1],))
                            rows = cursor1.fetchall()

                        amt = findrecord("gold_scheme_details", f"scheme_master_id = {scheme_masterid}", "sum(amount)")
                        gwt = findrecord("gold_scheme_details", f"scheme_master_id = {scheme_masterid}", "sum(gold_weight)")
                        if amt is None: amt = 0
                        if gwt is None: gwt = 0
                        for row1 in rows:
                            start_date = datetime.strptime(row1[1], "%d/%m/%Y")
                            next_due_date = start_date + relativedelta(months=1)
                            next_due_str = next_due_date.strftime("%Y-%m-%d")
                            cdate = datetime.strptime(row1[1], "%d/%m/%Y")
                            cdt = cdate.strftime("%Y-%m-%d")

                            # Insert into gold_scheme_details
                            insert_details_query = f"""
                                INSERT INTO gold_scheme_details 
                                (scheme_master_id, amount, gold_weight, gold_rate, status, payment_status, 
                                raz_order_id, transaction_id, next_due, `current_date`, payment_mode, payment_method)
                                VALUES ({scheme_masterid}, {row1[2]}, {row1[4]}, {row1[3]}, 1, 1, 'BANK', 
                                '{row1[5]}', '{next_due_str}', '{cdt}', 1, {row1[6]})
                            """
                            with connection.cursor() as cursor:
                                cursor.execute(insert_details_query)

                            amt += row1[2]
                            gwt += row1[4]

                            # Update gold_scheme_masters
                            update_query = f"""
                                UPDATE gold_scheme_masters 
                                SET amount_accumulated = {amt}, gold_accumulated = {int(gwt * 1000) / 1000}, 
                                total_gold_weight = {int(gwt * 1000) / 1000}, total_amount_accumulated = {amt} ,next_due='{next_due_str}'
                                WHERE erp_scheme_id = '{row[0]}' AND branch_id ={branchid}
                            """
                            with connection.cursor() as cursor1:
                                cursor1.execute(update_query)

                        # Update the status of the scheme

                        details = ""
                        try:
                            # Check if the row condition is met
                            if row[2] == 'N':
                                query = """
                                                           SELECT CAST(SALESMAST.Inv_Date AS VARCHAR(20)) + ' - ' + CAST(SALESMAST.Inv_No AS VARCHAR(10)) AS details
                                                           FROM SALECHITS
                                                           INNER JOIN SALESMAST ON SALECHITS.Serial_No = SALESMAST.Serial_No
                                                           WHERE SALECHITS.Chit_Key = %s
                                                       """
                                with connections[cn].cursor() as cursor1:
                                    cursor1.execute(query, [row[1]])  # Use parameterized query
                                    result = cursor1.fetchone()
                                    details = str(result[0]) if result else "No Details Found"
                            else:
                                if row[3] == 0:
                                    details = f"BANK-CHQ No. {row[5]}-{int(row[6])}"
                                else:
                                    details = "CASH"
                        except Exception as e:
                            # Handle any errors that might occur
                            details =""

                        update_status_query = """
                            UPDATE gold_scheme_masters
                            SET status = 3, closing_date = %s, invoice_number = %s
                            WHERE erp_scheme_id = %s AND branch_id = %s
                        """
                        date_obj = datetime.strptime(row[7], '%d/%m/%Y')
                        formatted_start_date =   date_obj.strftime("%Y-%m-%d")
                        with connection.cursor() as cursor:
                            cursor.execute(update_status_query, [ formatted_start_date, details, row[0], branchid])
                        MN=MN+1
                    # Return a success response after all rows are processed
                    return HttpResponse(f"""
                                    <!DOCTYPE html>
                                    <html lang="en">
                                    <head>
                                        <meta charset="UTF-8">
                                        <meta name="viewport" content="width=device-width, initial-scale=1.0">
                                        <title>Thank You</title>
                                        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
                                    </head>
                                    <body>
                                        <div class="container d-flex align-items-center justify-content-center vh-100">
                                            <div class="text-center">
                                                <div class="mb-4">
                                                    <img src="https://via.placeholder.com/150" alt="Thank You" class="img-fluid rounded-circle">
                                                </div>
                                                <h1 class="display-4 text-success">Thank You!</h1>
                                                <p class="lead">{MN} Schemes Closed successfully.</p>
                                                <p class="text-muted">We appreciate your trust in our services.</p>
                                                <a href="javascript:window.history.back();" class="btn btn-primary mt-3">Close</a>
                                            </div>
                                        </div>
                                    </body>
                                    </html>
                                """)
                else:
                    # If no records found, return a message
                    return HttpResponse("No records found for the given date range.", status=404)

        except Exception as e:
            # Log the error for debugging (e.g., use logging)
            print(e)
            return HttpResponse("An error occurred while processing your request. Please try again later." + e, status=500)
@login_required
def webupdateall(request):
    if request.POST:
        MN = 0
        branchid = request.user.branchid
        sdate = request.POST.get('start_date3', datetime.today().strftime('%m/%d/%Y'))
        edate = request.POST.get('end_date3', datetime.today().strftime('%m/%d/%Y'))

        # Parse dates based on their current format (assuming they might be in '%Y-%m-%d')
        try:
            sdate_obj = datetime.strptime(sdate, "%Y-%m-%d")  # Parse as YYYY-MM-DD
        except ValueError:
            sdate_obj = datetime.strptime(sdate, "%m/%d/%Y")  # Fallback to MM/DD/YYYY

        try:
            edate_obj = datetime.strptime(edate, "%Y-%m-%d")  # Parse as YYYY-MM-DD
        except ValueError:
            edate_obj = datetime.strptime(edate, "%m/%d/%Y")  # Fallback to MM/DD/YYYY

        # Format back to MM/DD/YYYY
        sdate = sdate_obj.strftime("%m/%d/%Y")
        edate = edate_obj.strftime("%m/%d/%Y")

        #try:
        cn = str(branchid)  # Assuming 'ekm' is the connection alias

        with connections[cn].cursor() as cursor1:

            query = """ 
                    IF NOT EXISTS (SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = 'CHITRCPTID')
                       CREATE TABLE CHITRCPTID (
                           id INT 
                       );"""

            cursor1.execute(query)
            # Fetch data from the database
            query = f"""
                       SELECT 
            cm.chit_key,
            cm.chit_no,
            cm.cust_name,
            cm.email,
            ccm.Address + ', ' + ccm.Address1 + ', ' + ccm.Address2 + ' [ Nominee :' + cm.Nominee + '( ' + cm.Nrelation + ' )]' AS Address, 
            CASE 
                WHEN cm.Mobile = '' THEN ccm.phone_no 
                ELSE cm.Mobile 
            END AS Phone_no,
            cm.Cust_Code,
            cm.inst_no,
            cm.doj
               FROM chitmast cm 
                JOIN CHITCUSTMAST ccm ON cm.ccust_key = ccm.CCust_key
                JOIN chitrcpt crp ON crp.chit_key = cm.chit_key
                WHERE cm.close_flag = 'N' 
                AND CONVERT(DATETIME, vch_date, 103) BETWEEN '{sdate}' AND '{edate}' and crp.serial_no not in (select id from CHITRCPTID where id is not null)
                ORDER BY cm.chit_key
                       """

            cursor1.execute(query)
            g=cursor1.fetchall()

            if g:

                        for ro in g:
                            cust_code=ro[6]
                            start_date = datetime.strptime(ro[8], "%d/%m/%Y")
                            next_due = start_date + relativedelta(months=1)
                            scheme_end_date = start_date + relativedelta(months=10)

                        # Format dates for SQL
                            start_date_str = start_date.strftime('%Y-%m-%d')
                            next_due_str = next_due.strftime('%Y-%m-%d')
                            scheme_end_date_str = scheme_end_date.strftime('%Y-%m-%d')

                            scheme_id = f"JS_O_{random.randint(0, 99999999):08d}"

                            # Check if the record exists in gold_scheme_masters
                            if not searchrecord("gold_scheme_masters",
                                            f"erp_scheme_id='{cust_code}' and branch_id={branchid}"):

                                # Insert into users table
                                query = """
                                INSERT INTO users (name, email, country_code, phone, status, is_complete) 
                                VALUES (%s, %s, '91', %s, 1, 1)
                                """
                                with connection.cursor() as cursor:
                                    cursor.execute(query, [ro[2], ro[3], ro[5]])

                                user_id = findrecord("users", f"phone='{ro[5]}' and name='{ro[2]}'", "id")

                                # Insert into user_kycs table
                                query = """
                                INSERT INTO user_kycs (user_id, name, address, type, status) 
                                VALUES (%s, %s, %s, 'Aadhar Card', 1)
                                """
                                with connection.cursor() as cursor:
                                    cursor.execute(query, [user_id, ro[2], ro[4]])

                                # Insert into gold_scheme_masters
                                query = f"""
                                INSERT INTO gold_scheme_masters 
                                (user_id, scheme_id, amount_accumulated, gold_accumulated, last_reduced, total_gold_weight, order_id, total_amount_accumulated, purchased_amount, branch_id, refund, discount_percentage, status, scheme_peroid, start_date, next_due, scheme_end_date, erp_scheme_id) 
                                VALUES ({user_id}, '{scheme_id}', 0, 0, 0, 0, 0, 0, 0, {branchid}, 0, 3.5, 1, {ro[7]}, '{start_date_str}', '{next_due_str}', '{scheme_end_date_str}', '{cust_code}')                    """

                                with connection.cursor() as cursor:
                                    cursor.execute(query)

                                # Find scheme master id
                                scheme_masterid = findrecord("gold_scheme_masters",
                                                         f"erp_scheme_id='{cust_code}' and branch_id={branchid}", "id")

                                # Insert into gold_scheme_details for each record in chitrcpt
                                query = """
                                SELECT 
                                    vch_no, vch_date, Amount, conv_rate, gold_wt, rcpt_no,
                                    CASE 
                                        WHEN trx_type = 'C' THEN 1 
                                         WHEN trx_type = 'A' THEN 
                                        CASE 
                                            WHEN cheque_no = '' THEN 2 
                                            ELSE 4 
                                        END
                                    WHEN trx_type = 'B' THEN 4 
                                    ELSE 4 
                                    END AS Trx, serial_no 
                                    FROM chitrcpt 
                                    WHERE chit_key = %s
                                """
                                with connections[cn].cursor() as cursor2:
                                    cursor2.execute(query, (ro[0],))
                                    rows = cursor2.fetchall()
                                amt = 0
                                gwt = 0

                                # Insert the records into gold_scheme_details
                                with connection.cursor() as cursor:
                                    for row1 in rows:
                                        start_date = datetime.strptime(row1[1], "%d/%m/%Y")
                                        next_due_date = start_date + relativedelta(months=1)
                                        next_due_str = next_due_date.strftime("%Y-%m-%d")

                                        cdate = datetime.strptime(row1[1], "%d/%m/%Y")

                                        cdt = cdate.strftime("%Y-%m-%d")

                                        query = f"""
                                        INSERT INTO gold_scheme_details 
                                        (scheme_master_id, amount, gold_weight, gold_rate, status, payment_status, raz_order_id, transaction_id, next_due, `current_date`, payment_mode, payment_method) 
                                        VALUES ({scheme_masterid}, {row1[2]}, {row1[4]}, {row1[3]}, 1, 1, 'BANK','{row1[5]}', '{next_due_str}', '{cdt}',  1, {row1[6]})
                                        """
                                        cursor.execute(query)

                                        query = f""" select id from CHITRCPTID where id={row1[7]} """
                                        with connections[cn].cursor() as cursor1:
                                            cursor1.execute(query)
                                            result = cursor1.fetchone()
                                        if not result:
                                            query = f"""insert into  CHITRCPTID (id) values( {row1[7]})  """

                                            with connections[cn].cursor() as cursor1:
                                                cursor1.execute(query)

                                            amt += row1[2]
                                            gwt += row1[4]

                                # Update gold_scheme_masters with accumulated amounts
                                query = f"""
                                UPDATE gold_scheme_masters 
                                SET amount_accumulated = {amt}, gold_accumulated = {gwt}, total_gold_weight = {gwt}, total_amount_accumulated = {amt},next_due='{next_due_str}'
                                WHERE erp_scheme_id = '{cust_code}' AND branch_id = {branchid}
                                """
                                with connection.cursor() as cursor:
                                    cursor.execute(query)

                            else:
                                # Existing scheme master
                                scheme_masterid = findrecord("gold_scheme_masters",
                                                         f"erp_scheme_id='{cust_code}' and branch_id={branchid}", "id")

                            # Delete previous details
                                query = "DELETE FROM gold_scheme_details WHERE scheme_master_id = %s and raz_order_id NOT LIKE %s and transaction_id NOT LIKE %s"
                                with connection.cursor() as cursor:
                                    cursor.execute(query, (scheme_masterid, 'order_%', 'pay_%'))

                                # Insert new details
                                query = """
                                SELECT 
                        vch_no, vch_date, chitrcpt.Amount, chitrcpt.conv_rate, gold_wt, chitrcpt.rcpt_no,
                        CASE 
                            WHEN chitrcpt.trx_type = 'C' THEN 1 
                            WHEN chitrcpt.trx_type = 'A' THEN 
                                CASE 
                                    WHEN cheque_no = '' THEN 2 
                                    ELSE 4 
                                END
                            WHEN chitrcpt.trx_type = 'B' THEN 4 
                            ELSE 4 
                        END AS Trx ,chitrcpt.serial_no
                    FROM chitrcpt,MCHITRCPTMAST 
                    WHERE chit_key = %s and chitrcpt.MSer_No =MCHITRCPTMAST.Serial_No and rcpttype=0 
                                """
                                with connections[cn].cursor() as cursor1:
                                    cursor1.execute(query, (ro[0],))
                                    rows = cursor1.fetchall()
                                amt = findrecord("gold_scheme_details", f"scheme_master_id = {scheme_masterid}",
                                                 "sum(amount)")
                                gwt = findrecord("gold_scheme_details", f"scheme_master_id = {scheme_masterid}",
                                                 "sum(gold_weight)")
                                if amt is None: amt = 0
                                if gwt is None: gwt = 0

                                # Insert the records into gold_scheme_details
                                with connection.cursor() as cursor:
                                    for row1 in rows:
                                        start_date = datetime.strptime(row1[1], "%d/%m/%Y")
                                        next_due_date = start_date + relativedelta(months=1)
                                        next_due_str = next_due_date.strftime("%Y-%m-%d")

                                        cdate = datetime.strptime(row1[1], "%d/%m/%Y")

                                        cdt = cdate.strftime("%Y-%m-%d")

                                        query = f"""
                                                               INSERT INTO gold_scheme_details 
                                                               (scheme_master_id, amount, gold_weight, gold_rate, status, payment_status, raz_order_id, transaction_id, next_due, `current_date`, payment_mode, payment_method) 
                                                               VALUES ({scheme_masterid}, {row1[2]}, {row1[4]}, {row1[3]}, 1, 1, 'BANK','{row1[5]}', '{next_due_str}', '{cdt}',  1, {row1[6]})
                                                               """
                                        cursor.execute(query)

                                        # Update CHITRCPTID

                                        query = f"""select id from CHITRCPTID where id={row1[7]} """
                                        with connections[cn].cursor() as cursor1:
                                            cursor1.execute(query)
                                            result = cursor1.fetchone()
                                        if not result:
                                            query = f"""insert into  CHITRCPTID (id) values( {row1[7]})  """

                                            with connections[cn].cursor() as cursor1:
                                                cursor1.execute(query)

                                            amt += row1[2]
                                            gwt += row1[4]

                                            # Update gold_scheme_masters with accumulated amounts
                                            query = f"""
                                                       UPDATE gold_scheme_masters 
                                                       SET amount_accumulated = {amt}, gold_accumulated ={gwt}, total_gold_weight = {gwt}, total_amount_accumulated = {amt} , next_due='{next_due_str}'
                                                       WHERE erp_scheme_id ='{cust_code}' AND branch_id = {branchid}
                                                       """
                                            with connection.cursor() as cursor1:
                                                cursor1.execute(query)
                        MN=MN+1
            return HttpResponse(f"""
                <!DOCTYPE html>
                <html lang="en">
                <head>
                    <meta charset="UTF-8">
                    <meta name="viewport" content="width=device-width, initial-scale=1.0">
                    <title>Thank You</title>
                    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
                </head>
                <body>
                    <div class="container d-flex align-items-center justify-content-center vh-100">
                        <div class="text-center">
                            <div class="mb-4">
                                <img src="https://via.placeholder.com/150" alt="Thank You" class="img-fluid rounded-circle">
                            </div>
                            <h1 class="display-4 text-success">Thank You!</h1>
                            <p class="lead">{MN} Schemes Updated or Inserted successfully.</p>
                            <p class="text-muted">We appreciate your trust in our services.</p>
                            <a href="javascript:window.history.back();" class="btn btn-primary mt-3">Close</a>
                        </div>
                    </div>
                </body>
                </html>
            """)
        # except Exception as e:
        #     print(f"Connection failed: {e}")
        # # Ensure that the view returns an HttpResponse even in case of an exception
        #     return HttpResponse("An error occurred while fetching data." + str(e))

@login_required
def chitsummary(request):
    if request.method == "POST":
        # Handle POST request as before
        start_date1 = request.POST.get('start_date1', date.today().strftime('%Y-%m-%d'))
        sdate = request.POST.get('start_date1', date.today().strftime('%d-%m-%Y'))
        end_date1 = request.POST.get('end_date1', date.today().strftime('%Y-%m-%d'))
        edate = request.POST.get('end_date1', date.today().strftime('%d-%m-%Y'))

        query = f"""
        SELECT 
            b.name AS BranchName,
            COALESCE(j.JoiningCount, 0) AS JoiningCount,
            COALESCE(j.JoiningAmount, 0) AS JoiningAmount,
            COALESCE(c.ClosingCount, 0) AS ClosingCount,
            COALESCE(c.ClosingAmount, 0) AS ClosingAmount,
            COALESCE(p.PaymentCount, 0) AS PaymentCount,
            COALESCE(p.PaymentAmount, 0) AS PaymentAmount,
            b.id
        FROM (
            SELECT DISTINCT gsm.branch_id 
            FROM gold_scheme_masters gsm
            UNION
            SELECT DISTINCT gsm.branch_id 
            FROM gold_scheme_masters gsm
            WHERE closing_date BETWEEN '2024-12-12' AND '2024-12-14'
            UNION
            SELECT DISTINCT gsm.branch_id 
            FROM gold_scheme_masters gsm
            JOIN gold_scheme_details gsd ON gsm.id = gsd.scheme_master_id
            WHERE `current_date` BETWEEN '{start_date1}' AND '{end_date1}'
        ) AS branches
        LEFT JOIN (
            SELECT gsm.branch_id, COUNT(*) AS JoiningCount, SUM(amount_accumulated) AS JoiningAmount
            FROM gold_scheme_masters gsm
            WHERE start_date BETWEEN '{start_date1}' AND '{end_date1}'
            GROUP BY gsm.branch_id
        ) AS j ON branches.branch_id = j.branch_id
        LEFT JOIN (
            SELECT gsm.branch_id, COUNT(*) AS ClosingCount, SUM(amount_accumulated) AS ClosingAmount
            FROM gold_scheme_masters gsm
            WHERE closing_date BETWEEN '{start_date1}' AND '{end_date1}'
            GROUP BY gsm.branch_id
        ) AS c ON branches.branch_id = c.branch_id
        LEFT JOIN (
            SELECT gsm.branch_id, COUNT(*) AS PaymentCount, SUM(gsd.amount) AS PaymentAmount
            FROM gold_scheme_masters gsm
            JOIN gold_scheme_details gsd ON gsm.id = gsd.scheme_master_id
            WHERE `current_date` BETWEEN '{start_date1}' AND '{end_date1}'
            GROUP BY gsm.branch_id
        ) AS p ON branches.branch_id = p.branch_id
        LEFT JOIN branches b ON branches.branch_id = b.id  order by j.JoiningCount desc;
        """
        with connection.cursor() as cursor1:
            cursor1.execute(query)
            rows = cursor1.fetchall()
            # Print the number of rows instead of using .count()

            data = [
                {
                    'branchname': row[0],
                    'jcount': row[1],
                    'jamount': int(row[2]),
                    'ccount': row[3],
                    'camount': int(row[4]),
                    'pcount': row[5],
                    'pamount': int(row[6]),
                    'b_id': row[7],
                }
                for row in rows  # Fix: closing parenthesis added for the for loop
            ]

        context = {'sdata': data,'sdate':start_date1,'edate':end_date1}
        # You can check the data printed here
        return render(request, "chitsummary.html", context)

    # Handle GET request
    return render(request, "chitsummary.html")


def chitreconf(request):
    if request.POST:
        branchid = request.user.branchid
        cust_code=request.POST.get("id")
        if cust_code=='':
            return HttpResponse("Please Enter Valid Scheme ID")

        cn = str(branchid)  # Assuming 'ekm' is the connection alias

        query = """
        SELECT 
            cm.chit_key,
            cm.chit_no,
            cm.cust_name,
            cm.email,
            ccm.Address + ', ' + ccm.Address1 + ', ' + ccm.Address2 + ' [ Nominee :' + cm.Nominee + '( ' + cm.Nrelation + ' )]' AS Address, 
            CASE 
                WHEN cm.Mobile = '' THEN ccm.phone_no 
                ELSE cm.Mobile 
            END AS Phone_no,
            cm.Cust_Code,
            cm.inst_no,
            cm.doj
        FROM chitmast cm 
        INNER JOIN CHITCUSTMAST ccm ON cm.ccust_key = ccm.CCust_key 
        WHERE cm.cust_code = %s
        """

        with connections[cn].cursor() as cursor1:
            cursor1.execute(query, [cust_code])  # Use parameterized query to prevent SQL injection
            result = cursor1.fetchone()

            if result:
                start_date = datetime.strptime(result[8], "%d/%m/%Y")
                next_due = start_date + relativedelta(months=1)
                scheme_end_date = start_date + relativedelta(months=10)

                # Format dates for SQL
                start_date_str = start_date.strftime('%Y-%m-%d')
                next_due_str = next_due.strftime('%Y-%m-%d')
                scheme_end_date_str = scheme_end_date.strftime('%Y-%m-%d')

                scheme_id = f"JS_O_{random.randint(0, 99999999):08d}"

                # Check if the record exists in gold_scheme_masters
                if not searchrecord("gold_scheme_masters", f"erp_scheme_id='{cust_code}' and branch_id={branchid}"):

                    # Insert into users table
                    query = """
                    INSERT INTO users (name, email, country_code, phone, status, is_complete) 
                    VALUES (%s, %s, '91', %s, 1, 1)
                    """
                    with connection.cursor() as cursor:
                        cursor.execute(query, [result[2], result[3], result[5]])

                    user_id = findrecord("users", f"phone='{result[5]}' and name='{result[2]}'", "id")

                    # Insert into user_kycs table
                    query = """
                    INSERT INTO user_kycs (user_id, name, address, type, status) 
                    VALUES (%s, %s, %s, 'Aadhar Card', 1)
                    """
                    with connection.cursor() as cursor:
                        cursor.execute(query, [user_id, result[2], result[4]])

                    # Insert into gold_scheme_masters
                    query = f"""
                    INSERT INTO gold_scheme_masters 
                    (user_id, scheme_id, amount_accumulated, gold_accumulated, last_reduced, total_gold_weight, order_id, total_amount_accumulated, purchased_amount, branch_id, refund, discount_percentage, status, scheme_peroid, start_date, next_due, scheme_end_date, erp_scheme_id) 
                    VALUES ({user_id}, '{scheme_id}', 0, 0, 0, 0, 0, 0, 0, {branchid}, 0, 3.5, 1, {result[7]}, '{start_date_str}', '{next_due_str}', '{scheme_end_date_str}', '{cust_code}')                    """
                    with connection.cursor() as cursor:
                        cursor.execute(query)

                    # Find scheme master id
                    scheme_masterid = findrecord("gold_scheme_masters",
                                                 f"erp_scheme_id='{cust_code}' and branch_id={branchid}", "id")

                    # Insert into gold_scheme_details for each record in chitrcpt
                    query = """
                    SELECT 
                        vch_no, vch_date, Amount, conv_rate, gold_wt, rcpt_no,
                        CASE 
                            WHEN trx_type = 'C' THEN 1 
                            WHEN trx_type = 'A' THEN 
                                CASE 
                                    WHEN cheque_no = '' THEN 2 
                                    ELSE 4 
                                END
                            WHEN trx_type = 'B' THEN 4 
                            ELSE 4 
                        END AS Trx, serial_no 
                    FROM chitrcpt 
                    WHERE chit_key = %s
                    """
                    cursor1.execute(query, [result[0]])
                    rows = cursor1.fetchall()
                    amt = 0
                    gwt = 0

                    # Insert the records into gold_scheme_details
                    with connection.cursor() as cursor:
                        for row1 in rows:
                            start_date = datetime.strptime(row1[1], "%d/%m/%Y")
                            next_due_date = start_date + relativedelta(months=1)
                            next_due_str = next_due_date.strftime("%Y-%m-%d")

                            cdate = datetime.strptime(row1[1], "%d/%m/%Y")

                            cdt = cdate.strftime("%Y-%m-%d")

                            query = f"""
                            INSERT INTO gold_scheme_details 
                            (scheme_master_id, amount, gold_weight, gold_rate, status, payment_status, raz_order_id, transaction_id, next_due, `current_date`, payment_mode, payment_method) 
                            VALUES ({scheme_masterid}, {row1[2]}, {row1[4]}, {row1[3]}, 1, 1, 'BANK','{row1[5]}', '{next_due_str}', '{cdt}',  1, {row1[6]})
                            """
                            cursor.execute(query)

                            query = f""" select id from CHITRCPTID where id={row1[7]} """
                            with connections[cn].cursor() as cursor1:
                                cursor1.execute(query)
                                result = cursor1.fetchone()
                            if not result:
                                query = f"""insert into  CHITRCPTID (id) values( {row1[7]})  """

                            with connections[cn].cursor() as cursor1:
                                cursor1.execute(query)

                            amt += row1[2]
                            gwt += row1[4]

                    # Update gold_scheme_masters with accumulated amounts
                            query = f"""
                    UPDATE gold_scheme_masters 
                    SET amount_accumulated = {amt}, gold_accumulated = {int(gwt * 1000) / 1000}, total_gold_weight = {int(gwt * 1000) / 1000}, total_amount_accumulated = {amt} ,next_due='{next_due_str}'
                    WHERE erp_scheme_id ='{cust_code}' AND branch_id = {branchid}
                    """
                    with connection.cursor() as cursor:
                        cursor.execute(query)
                    return HttpResponse(f"""
                                                           <!DOCTYPE html>
                                                           <html lang="en">
                                                           <head>
                                                               <meta charset="UTF-8">
                                                               <meta name="viewport" content="width=device-width, initial-scale=1.0">
                                                               <title>Thank You</title>
                                                               <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
                                                           </head>
                                                           <body>
                                                               <div class="container d-flex align-items-center justify-content-center vh-100">
                                                                   <div class="text-center">
                                                                       <div class="mb-4">
                                                                           <img src="https://via.placeholder.com/150" alt="Thank You" class="img-fluid rounded-circle">
                                                                       </div>
                                                                       <h1 class="display-4 text-success">Thank You!</h1>
                                                                       <p class="lead">{cust_code}  successfully updated.</p>
                                                                       <p class="text-muted">We appreciate your trust in our services.</p>
                                                                       <a href="javascript:window.history.back();" class="btn btn-primary mt-3">Close</a>
                                                                   </div>
                                                               </div>
                                                           </body>
                                                           </html>
                                                       """)
                else:
                    # Existing scheme master
                    scheme_masterid = findrecord("gold_scheme_masters",
                                                 f"erp_scheme_id='{cust_code}' and branch_id={branchid}", "id")

                    # Delete previous details
                    query = "DELETE FROM gold_scheme_details WHERE scheme_master_id = %s and raz_order_id NOT LIKE %s and transaction_id NOT LIKE %s"
                    with connection.cursor() as cursor:
                        cursor.execute(query, (scheme_masterid, 'order_%', 'pay_%'))

                    # Insert new details
                    query = """
                     SELECT 
                        vch_no, vch_date, chitrcpt.Amount, chitrcpt.conv_rate, gold_wt, chitrcpt.rcpt_no,
                        CASE 
                            WHEN chitrcpt.trx_type = 'C' THEN 1 
                            WHEN chitrcpt.trx_type = 'A' THEN 
                                CASE 
                                    WHEN cheque_no = '' THEN 2 
                                    ELSE 4 
                                END
                            WHEN chitrcpt.trx_type = 'B' THEN 4 
                            ELSE 4 
                        END AS Trx ,chitrcpt.serial_no
                    FROM chitrcpt,MCHITRCPTMAST 
                    WHERE chit_key = %s and chitrcpt.MSer_No =MCHITRCPTMAST.Serial_No and rcpttype=0 
                    """
                    cursor1.execute(query, [result[0]])
                    rows = cursor1.fetchall()
                    amt = findrecord("gold_scheme_details", f"scheme_master_id = {scheme_masterid}", "sum(amount)")
                    gwt = findrecord("gold_scheme_details", f"scheme_master_id = {scheme_masterid}", "sum(gold_weight)")
                    if amt is None: amt = 0
                    if gwt is None: gwt = 0
                    # Insert the records into gold_scheme_details
                    with connection.cursor() as cursor:
                        for row1 in rows:
                            start_date = datetime.strptime(row1[1], "%d/%m/%Y")
                            next_due_date = start_date + relativedelta(months=1)
                            next_due_str = next_due_date.strftime("%Y-%m-%d")

                            cdate = datetime.strptime(row1[1], "%d/%m/%Y")

                            cdt = cdate.strftime("%Y-%m-%d")

                            query = f"""
                                                       INSERT INTO gold_scheme_details 
                                                       (scheme_master_id, amount, gold_weight, gold_rate, status, payment_status, raz_order_id, transaction_id, next_due, `current_date`, payment_mode, payment_method) 
                                                       VALUES ({scheme_masterid}, {row1[2]}, {row1[4]}, {row1[3]}, 1, 1, 'BANK','{row1[5]}', '{next_due_str}', '{cdt}',  1, {row1[6]})
                                                       """
                            cursor.execute(query)

                            # Update CHITRCPTID

                            query = f"""select id from CHITRCPTID where id={row1[7]} """
                            with connections[cn].cursor() as cursor1:
                                cursor1.execute(query)
                                result = cursor1.fetchone()
                            if not result:
                                query = f"""insert into  CHITRCPTID (id) values( {row1[7]})  """

                            with connections[cn].cursor() as cursor1:
                                cursor1.execute(query)

                            amt += row1[2]
                            gwt += row1[4]

                            # Update gold_scheme_masters with accumulated amounts
                        query = f"""
                                            UPDATE gold_scheme_masters 
                                            SET amount_accumulated = {amt}, gold_accumulated = {int(gwt * 1000) / 1000}, total_gold_weight = {int(gwt * 1000) / 1000}, total_amount_accumulated = {amt} ,next_due='{next_due_str}'
                                            WHERE erp_scheme_id ='{cust_code}' AND branch_id = {branchid}
                                            """

                        with connection.cursor() as cursor:
                            cursor.execute(query)

                        return HttpResponse(f"""
                                        <!DOCTYPE html>
                                        <html lang="en">
                                        <head>
                                            <meta charset="UTF-8">
                                            <meta name="viewport" content="width=device-width, initial-scale=1.0">
                                            <title>Thank You</title>
                                            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
                                        </head>
                                        <body>
                                            <div class="container d-flex align-items-center justify-content-center vh-100">
                                                <div class="text-center">
                                                    <div class="mb-4">
                                                        <img src="https://via.placeholder.com/150" alt="Thank You" class="img-fluid rounded-circle">
                                                    </div>
                                                    <h1 class="display-4 text-success">Thank You!</h1>
                                                    <p class="lead">{cust_code}  updated successfully for Today.</p>
                                                    <p class="text-muted">We appreciate your trust in our services.</p>
                                                    <a href="javascript:window.history.back(); " class="btn btn-primary mt-3">Close</a>
                                                </div>
                                            </div>
                                        </body>
                                        </html>
                                    """)
            else:
                return HttpResponse("ID Not found")

def reconsil(request):
    if request.POST:
        branchid = request.user.branchid
        cn = str(branchid)  # Assuming branch ID is used as a connection identifier

        # Query 1
        query1 = """
        SELECT cust_code, cust_name, SUM(amount) AS Amount 
        FROM chitmast, chitrcpt   
        WHERE close_flag='N' AND chitrcpt.Chit_Key = chitmast.Chit_Key 
        GROUP BY cust_code, cust_name 
        ORDER BY cust_code
        """
        with connections[cn].cursor() as cursor1:
            cursor1.execute(query1)
            result = cursor1.fetchall()

        query5 = """
                SELECT  SUM(amount)-sum(taxable_amount) AS Amount 
                FROM chitmast, chitrcpt   
                WHERE close_flag='N' AND chitrcpt.Chit_Key = chitmast.Chit_Key 
                
                """
        with connections[cn].cursor() as cursor5:
            cursor5.execute(query5)
            result5 = cursor5.fetchone()





        # Query 2
        query2 = f"""
        SELECT erp_scheme_id, amount_accumulated 
        FROM gold_scheme_masters 
        WHERE branch_id = {branchid} 
        AND (closing_date IS NULL)
        ORDER BY erp_scheme_id;
        """
        with connection.cursor() as cursor2:
            cursor2.execute(query2)
            result1 = cursor2.fetchall()

        # Convert results into dictionaries for comparison
        result_dict = {row[0]: int(float(row[2])) for row in result if row[2]}  # cust_code: Amount
        result1_dict = {row[0]: int(float(row[1])) for row in result1 if
                        row[1] != 0}  # erp_scheme_id: amount_accumulated

        # Calculate the sums
        result_sum = sum(result_dict.values())  # Sum of amounts from result_dict
        result1_sum = sum(result1_dict.values())  # Sum of amounts from result1_dict

        # Count the entries
        resultcount = len(result_dict)
        result1count = len(result1_dict)

        # Find entries in result that are not in result1
        result_not_in_result1 = [
            (key, result_dict[key], "N/A") for key in result_dict if key not in result1_dict
        ]

        # Find entries in result1 that are not in result
        result1_not_in_result = [
            (key, "N/A", result1_dict[key]) for key in result1_dict if key not in result_dict
        ]

        result_result1 = [
            (key, result_dict[key], "N/A") for key in result_dict
        ]

        # Find entries in result1 that are not in result
        result1_in_result = [
            (key, "N/A", result1_dict[key]) for key in result1_dict
        ]

        # Combine unmatched entries
        unmatched_data = result_not_in_result1 + result1_not_in_result
        all_data = result_result1 + result1_in_result

        wb = Workbook()
        ws = wb.active
        ws.title = "Combined Data"

        # Write headers
        headers = ["Key", "Result Dict Value", "Result1 Dict Value"]
        ws.append(headers)

        # Write data
        for row in all_data:
            ws.append(row)

        # Save the Excel file
        file_path = "combined_data.xlsx"
        wb.save(file_path)
        print(file_path)
        # Pass unmatched data to the template
        context = {
            'unmatched_data': unmatched_data,
            "indsoftsum": float(result_sum)-result5[0],
            "onlinesum": float(result1_sum)-result5[0],
            "indsoftcount": resultcount,
            "onlinecount": result1count,
            "diff":result_sum-result1_sum,
            "alldata": all_data

        }

        return render(request, "reconsil.html", context)


from openpyxl.styles import Font
def indsoftdownload(request):
    if request.POST:
        branchid = request.user.branchid
        cn = str(branchid)  # Assuming branch ID is used as a connection identifier

        # Query 1
        query1 = """
        SELECT cust_code, cust_name, SUM(taxable_amount) AS Amount 
        FROM chitmast, chitrcpt   
        WHERE close_flag='N' AND chitrcpt.Chit_Key = chitmast.Chit_Key 
        GROUP BY cust_code, cust_name 
        ORDER BY cust_code
        """
        with connections[cn].cursor() as cursor1:
            cursor1.execute(query1)
            rows = cursor1.fetchall()
            query2 = f"""
                    SELECT erp_scheme_id, amount_accumulated 
                    FROM gold_scheme_masters 
                    WHERE branch_id = {branchid} 
                    AND (closing_date IS NULL)
                    ORDER BY erp_scheme_id;
                    """
            with connection.cursor() as cursor2:
                cursor2.execute(query2)
                result1 = cursor2.fetchall()

            # Convert results into dictionaries for comparison
            result_dict = {row[0]: int(float(row[2])) for row in rows  if row[2]}  # cust_code: Amount
            result1_dict = {row[0]: int(float(row[1])) for row in result1 if
                            row[1] != 0}  # erp_scheme_id: amount_accumulated

            combined_dict_overwrite = {**result_dict, **result1_dict}
            combined_dict_keep_both = {}

            for key in set(result_dict.keys()).union(result1_dict.keys()):
                combined_dict_keep_both[key] = (
                    result_dict.get(key, "N/A"),
                    result1_dict.get(key, "N/A")
                )

            # Decide which combined dictionary to export (e.g., overwrite version)
            combined_data = combined_dict_keep_both

            # Create Excel workbook
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Exported Data"

            # Write headers
            headers = ["Key", "Indsoft", "Online"]
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)

            # Write data rows
            for row_num, (key, values) in enumerate(combined_data.items(), start=2):
                sheet.cell(row=row_num, column=1, value=key)
                sheet.cell(row=row_num, column=2, value=values[0])  # Value from result_dict
                sheet.cell(row=row_num, column=3, value=values[1])  # Value from result1_dict

            # Prepare HTTP response for Excel download
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = 'attachment; filename="combined_data.xlsx"'

            # Save the workbook to the response
            workbook.save(response)
            return response
def updatemismatch(request):
    if request.POST:
        branchid = request.user.branchid
        cn = str(branchid)  # Assuming branch ID is used as a connection identifier

        # Query 1
        query1 = f"""WITH CTE_SchemeDetails AS (
    SELECT 
        scheme_master_id,
        SUM(amount) AS TotalAmount,
        TRUNCATE(SUM(gold_weight), 3) AS GoldWt
    FROM 
        gold_scheme_details
    GROUP BY 
        scheme_master_id
)
UPDATE 
    gold_scheme_masters gm
JOIN 
    CTE_SchemeDetails sd
    ON gm.id = sd.scheme_master_id
SET 
    gm.amount_accumulated = sd.TotalAmount,
    gm.gold_accumulated = sd.GoldWt
WHERE 
    gm.branch_id = {branchid}
    AND (gm.amount_accumulated <> sd.TotalAmount 
         OR gm.gold_accumulated <> sd.GoldWt);
           """
        with connection.cursor() as cursor1:
            cursor1.execute(query1)
            affected_rows = cursor1.rowcount
    return HttpResponse( str(affected_rows) + " Rows affected")





import razorpay
import requests
from django.conf import settings
from django.http import JsonResponse
RAZ_SECRET_KEY = "uY65nGUy3Ur3m7m7CcxtTsur"
RAZ_API_KEY = "rzp_live_3UTNfDkLZJN3nj"
# Initialize Razorpay client
client = razorpay.Client(auth=(RAZ_API_KEY, RAZ_SECRET_KEY))

def fetch_orders(orderid):
    orders = client.order.all()  # Fetch all orders
    items = orders.get('items', [])  # Extract items list

    # Find and return the details of the specific order
    specific_order = next((item for item in items if item.get('id') == orderid), None)
    return specific_order

    # Fetch payments
def fetch_payments( orderid):
        payments = client.payment.all()
        items = payments.get('items', [])  # Extract items list
        specific_order = next((item for item in items if item.get('order_id') == orderid), None)
        return specific_order

    # Fetch settlements
def fetch_settlement_by_order( order_id):
    url = "https://api.razorpay.com/v1/settlements/recon/combined?year=2025&month=01"

    # Set up authentication and headers
    auth = HTTPBasicAuth(RAZ_API_KEY , RAZ_SECRET_KEY)
    headers = {
        'Content-Type': 'application/json'
    }

    try:
        # Send GET request to Razorpay API
        response = requests.get(url, auth=auth, headers=headers)

        # Check if the response is successful
        response.raise_for_status()  # Will raise an error if status code is 4xx or 5xx

        # Parse the JSON response
        data = response.json()

        # Filter settlements by order_id
        filtered_data = [
            entry for entry in data.get("items", [])
            if entry.get("order_id") == order_id
        ]

        if filtered_data:
            return filtered_data
        else:
            return f"No settlements found for order_id: {order_id}"

    except requests.exceptions.HTTPError as http_err:
        return f"HTTP error occurred: {http_err}"
    except requests.exceptions.RequestException as err:
        return f"Error: {err}"


# Combine Orders, Payments, and Settlements
# def combine_data(orders, payments, settlements):
#         combined_data = []
#
#         # Map payments by order_id for quick lookup
#         payment_map = {payment['order_id']: payment for payment in payments if payment.get('order_id')}
#
#         # Map settlements by UTR for quick lookup
#         settlement_map = {settlement['utr']: settlement for settlement in settlements if settlement.get('utr')}
#
#         # Combine data
#         for order in orders:
#             order_id = order.get('id', 'N/A')
#             payment = payment_map.get(order_id, {})
#             utr = payment.get('utr', 'N/A')
#             settlement = settlement_map.get(utr, {})
#
#             # Combine order, payment, and settlement details
#             combined_record = {
#                 "order_id": order_id,
#                 "order_amount": order.get('amount', 0)/100,
#                 "order_status": order.get('status', 'N/A'),
#                 "payment_id": payment.get('id', 'N/A'),
#                 "payment_amount": payment.get('amount', 0),
#                 "payment_status": payment.get('status', 'N/A'),
#                 "utr": utr,
#                 "settlement_id": settlement.get('id', 'N/A'),
#                 "settlement_status": settlement.get('status', 'N/A'),
#                 "settlement_amount": settlement.get('amount', 0)
#             }
#
#             combined_data.append(combined_record)
#             print(combined_data)
#         return (combined_data)

import json
def razorpayreport(request):
    year, month, day = 2025, 1, 2
    order_id = "order_PaYHNTmMzjx69H"

    result = fetch_settlement_by_order("", "api_secret",order_id )
    print(result)

    # combined_data = combine_data(orders, payments, settlements)
    # for order in combined_data:
    #     print(order["order_id"])



    # combined_json = json.dumps(combined_data, indent=4)
    #
    # return render(request,"razorpayreport.html",{"data":combined_data})
    #
    return HttpResponse("Nothing")
